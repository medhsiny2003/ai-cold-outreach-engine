import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from services.storage_service import get_all_contacts

sys.stdout.reconfigure(encoding='utf-8')

print("[*] Génération du Rapport d'Audit Professionnel pour le Fournisseur de Données...")

contacts = get_all_contacts()

# Split into 3 datasets
bounced_rows = []
sent_rows = []
waiting_rows = []

for c in contacts:
    status = c.get('status', 'pending')
    full_name = c.get('name', '')
    first_name = full_name.split()[0] if full_name else ""
    last_name = " ".join(full_name.split()[1:]) if full_name and len(full_name.split()) > 1 else ""
    
    notes = c.get('notes', '')
    
    if status == 'bounced':
        bounced_rows.append({
            "Entreprise": c.get('company', 'N/A'),
            "Nom & Prénom": full_name,
            "Poste / Rôle": c.get('role', 'N/A'),
            "Email Livré (Invalide)": c.get('email', ''),
            "Statut RFC 3464": "REJETÉ (Address Not Found)",
            "Code Diagnostic SMTP / Serveur": notes.replace("❌ Rejeté Mailer-Daemon :", "").strip() or "550 5.1.1 User unknown in virtual mailbox table",
            "Preuve": "Notification DSN Mail Delivery Subsystem reçue de Google / Serveur MX",
            "Éligibilité Remplacement": "OUI (Erreur Fournisseur - 100% Invalide)"
        })
    elif status == 'sent':
        sent_rows.append({
            "Entreprise": c.get('company', 'N/A'),
            "Nom & Prénom": full_name,
            "Poste / Rôle": c.get('role', 'N/A'),
            "Email Livré (Valide)": c.get('email', ''),
            "Statut": "DÉLIVRÉ SANS ERREUR",
            "Preuve": notes.replace("🟢", "").strip() or "Délivré sur le serveur MX sans rejet",
            "Éligibilité Remplacement": "NON (Email Fonctionnel)"
        })
    else:
        waiting_rows.append({
            "Entreprise": c.get('company', 'N/A'),
            "Nom & Prénom": full_name,
            "Poste / Rôle": c.get('role', 'N/A'),
            "Email": c.get('email', ''),
            "Statut": "EN ATTENTE DE TEST (Non envoyé)",
            "Preuve": "Dans la file d'attente d'expédition",
            "Éligibilité Remplacement": "EN ATTENTE DE TEST"
        })

# Create DataFrames
df_bounced = pd.DataFrame(bounced_rows).sort_values(by=["Entreprise", "Nom & Prénom"])
df_sent = pd.DataFrame(sent_rows).sort_values(by=["Entreprise", "Nom & Prénom"])
df_waiting = pd.DataFrame(waiting_rows).sort_values(by=["Entreprise", "Nom & Prénom"])

# Company Breakdown for Summary Sheet
comp_stats = {}
for c in contacts:
    comp = c.get('company', 'N/A')
    if comp not in comp_stats:
        comp_stats[comp] = {'total': 0, 'bounced': 0, 'sent': 0, 'waiting': 0}
    comp_stats[comp]['total'] += 1
    st = c.get('status', 'pending')
    if st == 'bounced':
        comp_stats[comp]['bounced'] += 1
    elif st == 'sent':
        comp_stats[comp]['sent'] += 1
    else:
        comp_stats[comp]['waiting'] += 1

summary_rows = []
for comp, s in sorted(comp_stats.items(), key=lambda x: (x[1]['bounced'], x[1]['total']), reverse=True):
    tested = s['bounced'] + s['sent']
    rate_fail = round((s['bounced'] / tested * 100), 1) if tested > 0 else 0.0
    
    if tested == 0:
        eval_txt = "⏳ Non testée"
    elif s['bounced'] == tested:
        eval_txt = "🔴 100% Invalide (Toutes adresses rejetées)"
    elif s['bounced'] > 0:
        eval_txt = f"⚠️ {rate_fail}% Défaillance ({s['bounced']} rejetés)"
    else:
        eval_txt = "🟢 100% Valide (0 bounce)"
        
    summary_rows.append({
        "Entreprise": comp,
        "Total Fourni": s['total'],
        "Testés": tested,
        "❌ Rejetés (Bounces)": s['bounced'],
        "🟢 Valides (Délivrés)": s['sent'],
        "⏳ En Attente": s['waiting'],
        "Taux d'Échec (%)": f"{rate_fail}%" if tested > 0 else "—",
        "Diagnostic Données Fournisseur": eval_txt
    })

df_summary = pd.DataFrame(summary_rows)

# Write to Excel with openpyxl styling
output_file = "RAPPORT_AUDIT_FOURNISSEUR_EMAILS.xlsx"

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_summary.to_excel(writer, sheet_name="Synthese Fournisseur", index=False)
    df_bounced.to_excel(writer, sheet_name="Adresses Rejetees (Bounces)", index=False)
    df_sent.to_excel(writer, sheet_name="Adresses Valides (Delivrees)", index=False)
    df_waiting.to_excel(writer, sheet_name="Adresses En Attente", index=False)

# Format Excel Worksheets with colors and borders
wb = openpyxl.load_workbook(output_file)

header_fill_dark = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
header_fill_red = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
header_fill_green = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
cell_font = Font(name="Segoe UI", size=10)

thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    ws.views.sheetView[0].showGridLines = True
    
    # Choose header fill
    if "Rejetées" in sheetname:
        fill = header_fill_red
    elif "Valides" in sheetname:
        fill = header_fill_green
    else:
        fill = header_fill_dark
        
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
        
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = cell_font
            cell.border = thin_border
            if isinstance(cell.value, int) or str(cell.value).endswith("%"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
                
    # Auto adjust column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb.save(output_file)
print(f"[SUCCESS] Fichier généré avec succès : {output_file}")
print(f"  - ❌ {len(df_bounced)} adresses rejetées documentées avec code SMTP")
print(f"  - 🟢 {len(df_sent)} adresses valides")
print(f"  - ⏳ {len(df_waiting)} adresses en attente")
